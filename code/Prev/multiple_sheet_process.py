import sys
sys.path.append('/home/cdsw/code')
from utils import * 
import math 
import dateutil
from dateutil.parser import parse
import openpyxl
import numpy as np
import os
import pandas as pd


def traiter_retraitement(filename, log):
    final = pd.DataFrame()
    sheetnames = openpyxl.load_workbook(filename).sheetnames 
    name = filename.split("Prévisions DEA mensualisées/")[1]
    for sheet in sheetnames[:5] :
        df = pd.read_excel(filename , sheet_name = sheet, engine = 'openpyxl')
        #print(filename, "   ", sheet)
        df = make_date_header(df)
        df= table_delimiter(df)

        atraiter = False 
        ## recherche de la colonne contenant la valeur retraitements et celle contenant la valeur raisons

        for column in df.columns : 
            try : 
                if df[column].str.contains('raison' , case = False).sum() >= 1: 
                    col_raison  = column
                    atraiter = True
            except : 
                pass

        if atraiter : 

            print("le fichier ", name, " et onglet ", sheet, " est à traiter")
            log.info(f" le fichier {name} de feuille {sheet} est à traiter")

            # recherche de la ligne contenant la valeur retraitements et celle contenant la valeur raisons
            row_retraitement = [i for i in range(len(df)) if df[col_raison].str.contains('Retraitement' ,
                                                                                               case = False).tolist()[i] == True]
            row_raison = [i for i in range(len(df)) if df[col_raison].str.contains('raison' ,
                                                                                         case = False).tolist()[i] == True]

            # construire le Dataframe contenant les retraitement et leurs raisons 
            result = pd.DataFrame()
            temp = pd.DataFrame()
            # ajouter les valeurs de retraitement et de raison 

            for i in range(len(row_raison)):
                temp["montant_retraitement"] = df.iloc[row_retraitement[i]]
                temp["raison"] = df.iloc[row_raison[i]]

                result = pd.concat([result, temp])

            result.dropna(subset = ['montant_retraitement'],inplace = True)


            result['nom_fichier'] = filename.split("Prévisions DEA mensualisées/")[1]
            result['nom_onglet'] = sheet

            # make first row as header

            result = result[1:]
            result = result.reset_index()    
            result = delete_noise(result,result.columns[0])  
            if len(result) == 0:
                print("mais ne contient pas de retraitements à traiter")
                log.info('mais ne contient pas de retraitements à traiter')
            result = result.rename(columns = { result.columns[0] : "date"})

            ## ajouter la date 
            result['year'] = pd.DatetimeIndex(result['date']).year
            result['month'] = pd.DatetimeIndex(result['date']).month
            result['month'] = result['month'].apply( lambda x : add_zero(x) )
            mapping_month = {'01': 'janvier' , '02':'février' , '03': 'mars', '04': 'avril', '05':'mai', '06': 'juin' , 
                             '07':'juillet', '08' : 'août' ,'09' : 'septembre' , '10' : 'octobre' , '11' : 'novembre' , 
                             '12' : 'décembre' }
            result['mois_nom'] = result['month'].map(mapping_month)

            result = result.rename(columns = {"year" : "annee", "month" : "mois"})


            result = result[['nom_fichier', 'nom_onglet', 'annee', 'mois', 'mois_nom' ] + 
                            [e for e in result.columns if 'retraitement' in e] + [e for e in result.columns if "raison" in e]]
            
            final = pd.concat([final, result])
        else:
            print("le fichier ",name, " et onglet ", sheet,
                 " ne contient pas de retraitements")
            log.info(f" le fichier {name} de feuille {sheet} ne contient pas de retraitements")

            

    return final.reset_index().drop(columns = "index")


def prevision_data_ingestion_multiple_sheet(spark, depot, db_name, table_name, filename, extension, normal_values_nature_montant, log):
    """
    The function is useful to create a pandas dataframe from a multisheets excel files
    params : 
    - filename : the name of the file to be treated
    - extension : the extension of the file to be treated   
    """
    log.info(f"Début d'écriture du fichier {filename.split('plateforme/')[1]}")
    
    startTime = datetime.now()
    dataframe = pd.DataFrame() 

    
    for annee in openpyxl.load_workbook(filename).sheetnames :
        if extension == '*.xlsx' :
            df = pd.read_excel(filename, sheet_name = annee ,engine = 'openpyxl')
        elif extension == '*.csv':
            df = pd.read_csv(filename, sheet_name = annee ,engine = 'openpyxl')
            
        #--------------------------------------------------------------------------------------------------------------#
        #-----supprimer les tableaux secondaires que nous ne souhaitons pas récupérer à partir du fichier excel/csv----#
        #--------------------------------------------------------------------------------------------------------------#
        stop_l, stop_c, column_bloc = check_index(df)
        if stop_c[0] != column_bloc: 
            df.loc[df[df.columns[stop_c[0]]].isna(), df.columns[stop_c[0]]] = \
            df[df[df.columns[stop_c[0]]].isna()][df.columns[column_bloc]]
        df =df[df.columns[stop_c[0]:stop_c[1]]]
        df =df[:stop_l]
        df = df.dropna(axis = 0, how = 'all')   ## supprimer les lignes de séparation 
        #--------------------------------------------------------------------------------------------------------------#
        #------------------------transposer les données et renommer la colonne mois------------------------------------#
        #--------------------------------------------------------------------------------------------------------------#
        
        df = df.T
        df= df.reset_index(level=0)
        df = df.rename(columns = {'index'  : "mois_nom"})

        #--------------------------------------------------------------------------------------------------------------#
        #---------------------------------------remplir la colonne année-----------------------------------------------#
        #--------------------------------------------------------------------------------------------------------------#

        df['annee'] = annee
        df = df[['annee','mois_nom',0] + df.columns[2:-1].tolist()]
        df.head()
        
        #--------------------------------------------------------------------------------------------------------------#
        #-----------------------------récupérer la première partie de la table déjà traitée----------------------------#
        #--------------------------------------------------------------------------------------------------------------#
        
        df1= df[df.columns[:3]]
        # changer l'entete de la table
        df1= df1[1:].reset_index().drop(columns = 'index')
        df1 = df1.rename(columns = {0:'type_montant'})
        # ajouter une colonne 'key' pour une jointure plus tard
        df1['key'] = [i for i in range(len(df1))]
        
        #-------------------------------------------------------------------------------------------------------------#
        #----------------------------transformer la deuxième partie de la table---------------------------------------#
        #-------------------------------------------------------------------------------------------------------------#
        
        df2= df[df.columns[3:]]
        new_header = df2.iloc[0] 
        df2 = df2[1:] 
        df2.columns = new_header 

        df3= df2.T.unstack().reset_index(level=1, name='montant_millions')
        df3= df3.rename(columns ={0:'nature_montant'})
        df3.loc[df3['nature_montant'].isna(), 'nature_montant'] = 'Eléments exceptionnels'

        l = [len(df2.columns)*[i] for i in range(len(df1))]
        flat_l = [item for sublist in l for item in sublist]

        df3['key'] = flat_l

        dff = df3.merge(df1, on= 'key')
        dff.drop(columns ='key', inplace = True)
        dff['montant_millions'] = dff['montant_millions'].astype(float, errors = 'ignore')
        
        #-----------------------------------------------------------------------------------------------------------#
        #-------------dictionnaire de mapping pour ajouter les mois sous formation 'int'----------------------------#
        #-----------------------------------------------------------------------------------------------------------#
        mapping_month = {'janvier': '01' , 'février':'02' , 'mars': '03', 'avril': '04', 'mai':'05', 'juin': '06' , 'juillet':'07' , 'août' : '08' , 'septembre' : '09' , 'octobre' : '10' , 'novembre' : '11' , 'décembre' : '12' }

        dff['mois'] = dff['mois_nom'].map(mapping_month)

        #-----------------------------------------------------------------------------------------------------------#
        #---------ajouter la colonne version_prev qui indique la date de génération du fichier de prévision---------#
        #-----------------------------------------------------------------------------------------------------------#
       
        
        dff['version_prev'] = get_date_extraction_m_yy(filename)
        dff['version_prev'] = pd.to_datetime(dff['version_prev'])
        dff['version_prev']= dff['version_prev'].dt.strftime('%Y-%m')

        # trim de la colonne nature_montant
        dff['nature_montant'] = dff['nature_montant'].str.strip()
      
            
        dff = dff[['version_prev','annee' , 'mois' , 'mois_nom', 
                                'nature_montant',  'type_montant', 'montant_millions' ]]
        dff = dff.replace('null', np.nan) 
        dff.dropna(subset = ['montant_millions' ] ,inplace = True )

        end_time = str(datetime.now() - startTime)
        print("durée de traitement = " + end_time)
        log.info("durée de traitement = " + end_time)  
        
        #-----------------------------------------------------------------------------------------------------------#
        #---------------------ajouter le dataframe s'il vérifie les critères de Data Quality------------------------#
        #-----------------------------------------------------------------------------------------------------------#

        if data_quality(dff, normal_values_nature_montant)[1]:        
            dataframe = pd.concat([dataframe, dff])
            log.info(f" le fichier {filename.split('plateforme/')[1]} de feuille {annee} respecte les critères de Data Quality et sera écrit")
        else:
            log.warning(f" le fichier {filename.split('plateforme/')[1]} de feuille {annee} ne sera pas écrit car ne respecte pas les critères de Data Quality")

    # écrire le dataframe final sur hdfs   
    load_aggregate_table_to_hdfs(spark,depot,db_name,table_name,dataframe,log)
    log.info(f"Fin d'écriture du fichier {filename.split('plateforme/')[1]} avec succès")
        
    return dataframe.reset_index(drop = True) 
